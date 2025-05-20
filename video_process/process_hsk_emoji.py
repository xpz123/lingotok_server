import pandas as pd
import os
os.environ["IMAGEIO_FFMPEG_EXE"] = "/opt/homebrew/Cellar/ffmpeg/7.1_4/bin/ffmpeg"
from translator import translate_text2ar
from pypinyin import pinyin

from moviepy.editor import VideoFileClip, AudioFileClip, TextClip, CompositeVideoClip, ColorClip, concatenate_videoclips, CompositeAudioClip, ImageSequenceClip
from moviepy.config import change_settings
change_settings({"IMAGEMAGICK_BINARY": "/opt/homebrew/Cellar/imagemagick/7.1.1-43/bin/magick"})
from bidi.algorithm import get_display
import arabic_reshaper
from llm_util import call_doubao_pro_32k
from huoshan_tts_util import generate_wav

from tqdm import tqdm

import subprocess
import re
import json
from retrying import retry
from video_processor import VideoProcessor, milliseconds_to_time_string
import time
from aigc import merge_audios
import random as rd
from copy import deepcopy

from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor


def generate_subtitle_singleword(zh_text_list, ar_text_list, en_text_list, srt_prefix, root_dir, word_dur):
    video_processor = VideoProcessor()
    res = dict()
    zh_srt = os.path.join(root_dir, "{}_Chinese.srt".format(srt_prefix))
    ar_srt = os.path.join(root_dir, "{}_Arabic.srt".format(srt_prefix))
    en_srt = os.path.join(root_dir, "{}_English.srt".format(srt_prefix))
    pinyin_srt = os.path.join(root_dir, "{}_Pinyin.srt".format(srt_prefix))

    res["zh_srt"] = zh_srt
    res["ar_srt"] = ar_srt
    res["en_srt"] = en_srt
    res["pinyin_srt"] = pinyin_srt
    
    # ar_text_list = translate_text2ar(zh_text_list, "ar")
    assert len(ar_text_list) == len(zh_text_list)

    # en_text_list = translate_text2ar(zh_text_list, "en")
    assert len(en_text_list) == len(zh_text_list)

    fw_zh = open(zh_srt, "w")
    fw_ar = open(ar_srt, "w")
    fw_en = open(en_srt, "w")
    
    start_time = 0
    for i in range(len(zh_text_list)):
        dur_ms = word_dur
        start_time_str = milliseconds_to_time_string(start_time)
        end_time_str = milliseconds_to_time_string(start_time + dur_ms)
        start_time = start_time + dur_ms
        zh_srt_content = f"{i}\n{start_time_str} --> {end_time_str}\n{zh_text_list[i]}\n\n"
        fw_zh.write(zh_srt_content)

        ar_srt_content = f"{i}\n{start_time_str} --> {end_time_str}\n{ar_text_list[i]}\n\n"
        fw_ar.write(ar_srt_content)

        en_srt_content = f"{i}\n{start_time_str} --> {end_time_str}\n{en_text_list[i]}\n\n"
        fw_en.write(en_srt_content)
    fw_zh.close()
    fw_ar.close()
    fw_en.close()

    video_processor.convert_zhsrt_to_pinyinsrt(zh_srt, pinyin_srt)
    return res

def trans_word_to_ar(ori_csv_file, ar_csv_file):
    df = pd.read_csv(ori_csv_file)
    columns = df.columns.to_list()
    columns.append("阿语翻译")
    text_list = list()
    for i in tqdm(range(df.shape[0])):
        text = df.iloc[i]["单词"]
        text_list.append(text.split("（")[0])
    resp_list = translate_text2ar(text_list, "ar")
    assert len(text_list) == len(resp_list)
    ar_text_list = [resp["Translation"] for resp in resp_list]
    df["阿语翻译"] = ar_text_list
    df.to_csv(ar_csv_file, index=False)

def trans_word_to_en(ori_csv_file, ar_csv_file):
    df = pd.read_csv(ori_csv_file)
    columns = df.columns.to_list()
    columns.append("英语翻译")
    text_list = []
    en_text_list = []
    
    # 收集所有需要翻译的文本
    for i in tqdm(range(df.shape[0])):
        text = df.iloc[i]["单词"]
        text_list.append(text.split("（")[0])
    
    # 按100个一组进行批量翻译
    batch_size = 100
    for i in tqdm(range(0, len(text_list), batch_size)):
        batch_texts = text_list[i:i + batch_size]
        try:
            resp_list = translate_text2ar(batch_texts, "en")
            assert len(batch_texts) == len(resp_list)
            batch_translations = [resp["Translation"] for resp in resp_list]
            en_text_list.extend(batch_translations)
        except Exception as e:
            en_text_list.extend([""] * len(batch_texts))
            time.sleep(5)
    
    assert len(text_list) == len(en_text_list)
    df["英语翻译"] = en_text_list
    df.to_csv(ar_csv_file, index=False)

def trans_word_to_en_with_mapcsv(ori_csv_file, en_csv_file):
    df = pd.read_excel("/Users/tal/work/lingtok_server/video_process/hsk_dictionary/中英单词.xlsx")
    zh_en_map = dict()
    for i in range(df.shape[0]):

        if df.iloc[i]["是否正确"] == "正确":
            zh_en_map[df.iloc[i]["单词"].strip()] = df.iloc[i]["英语翻译"].strip().lower()
        else:
            if not pd.isna(df.iloc[i]["修改结果"]):
                zh_en_map[df.iloc[i]["单词"].strip()] = df.iloc[i]["修改结果"].strip()
            else:
                continue
    import pdb;pdb.set_trace()

    df = pd.read_csv(ori_csv_file)
    columns = df.columns.to_list()
    columns.append("英语翻译")
    en_text_list = []
    
    for i in range(df.shape[0]):
        word = df.iloc[i]["原始单词"]
        # 从映射字典中查找英文翻译
        en_translation = zh_en_map.get(word, None)
        en_text_list.append(en_translation)
    
    df["英语翻译"] = en_text_list
    df.to_csv(en_csv_file, index=False)

def add_pinyin(ori_csv_file, py_csv_file):
    df = pd.read_csv(ori_csv_file)
    df.dropna(subset=["单词"], axis=0, how="any", inplace=True)
    columns = df.columns.to_list()
    columns.append("拼音")
    df_list = df.values.tolist()
    for i in range(df.shape[0]):
        py_list = pinyin(df.iloc[i]["单词"])
        py_str = ""
        for item in py_list:
            for py in item:
                py_str += " " + py
        df_list[i].append(py_str)
    df_new = pd.DataFrame(df_list, columns=columns)
    df_new.to_csv(py_csv_file, index=False)

def generate_subtitle(zh_text_list, srt_prefix, root_dir, audio_dur_dict, audio_list=None):
    video_processor = VideoProcessor()
    res = dict()
    zh_srt = os.path.join(root_dir, "{}_Chinese.srt".format(srt_prefix))
    ar_srt = os.path.join(root_dir, "{}_Arabic.srt".format(srt_prefix))
    en_srt = os.path.join(root_dir, "{}_English.srt".format(srt_prefix))
    pinyin_srt = os.path.join(root_dir, "{}_Pinyin.srt".format(srt_prefix))

    res["zh_srt"] = zh_srt
    res["ar_srt"] = ar_srt
    res["en_srt"] = en_srt
    res["pinyin_srt"] = pinyin_srt
    
    ar_text_list = translate_text2ar(zh_text_list, "ar")
    assert len(ar_text_list) == len(zh_text_list)

    en_text_list = translate_text2ar(zh_text_list, "en")
    assert len(en_text_list) == len(zh_text_list)

    fw_zh = open(zh_srt, "w")
    fw_ar = open(ar_srt, "w")
    fw_en = open(en_srt, "w")
    
    start_time = 0
    for i in range(len(zh_text_list)):
        if audio_list is not None:
            dur_ms = int(audio_dur_dict[audio_list[i]] * 1000)
        else:
            dur_ms = int(audio_dur_dict["{}_{}".format(zh_text_list[0], (i+1))] * 1000)
        start_time_str = milliseconds_to_time_string(start_time)
        end_time_str = milliseconds_to_time_string(start_time + dur_ms)
        start_time = start_time + dur_ms
        zh_srt_content = f"{i}\n{start_time_str} --> {end_time_str}\n{zh_text_list[i]}\n\n"
        fw_zh.write(zh_srt_content)

        ar_srt_content = f"{i}\n{start_time_str} --> {end_time_str}\n{ar_text_list[i]['Translation']}\n\n"
        fw_ar.write(ar_srt_content)

        en_srt_content = f"{i}\n{start_time_str} --> {end_time_str}\n{en_text_list[i]['Translation']}\n\n"
        fw_en.write(en_srt_content)
    fw_zh.close()
    fw_ar.close()
    fw_en.close()

    video_processor.convert_zhsrt_to_pinyinsrt(zh_srt, pinyin_srt)
    return res

def generate_trans_quiz(en_ar_csv, quiz_csv, quiz_jsonl):
    def get_distractor_words(word, content_tagger):
        level_idx = {"一级": 1, "二级": 2, "三级": 3, "四级": 4, "五级": 5, "六级": 6}
        word_level = word.split("（")[-1].replace("）", "").strip()
        if word_level in level_idx:
            level = level_idx[word_level]
        else:
            level = 6
        word_list = content_tagger.get_hsk_level_word_list(level)
        rd.shuffle(word_list)
        distractor_words = word_list[:3]
        return distractor_words
    
    quiz_template = {"sentence": "", "question": "", "options": [], "answer": "", "explanation": "", "ar_question": "", "ar_options": [], "ar_explanation": "", "en_question": "", "en_options": [], "en_explanation": "", "vid": ""}
    from content_tagger import ContentTagger
    content_tagger = ContentTagger()

    df = pd.read_csv(en_ar_csv)
    columns = df.columns.to_list()
    columns.append("问题")
    columns.append("quiz_id")
    df_list = df.values.tolist()
    fw = open(quiz_jsonl, "w", encoding="utf-8")

    for i in tqdm(range(df.shape[0])):
        word = df.iloc[i]["单词"]
        ori_word = df.iloc[i]["原始单词"]
        vid_id = "{}_{}".format(i, ori_word)
        ar_word = df.iloc[i]["阿语翻译"]
        en_word = df.iloc[i]["英语翻译"]
        if (not pd.isna(en_word)) and  (not pd.isna(ar_word)):
            # 获取干扰选项单词
            content = deepcopy(quiz_template)
            distractor_words = get_distractor_words(ori_word, content_tagger)
            options = distractor_words
            options.append(word)
            rd.shuffle(options)
            idx2alp = ["A", "B", "C", "D"]
            for opt_idx, opt in enumerate(options):
                if opt == word:
                    content["answer"] = idx2alp[opt_idx]
                    break
            options = [idx2alp[i] + ". " + options[i] for i in range(len(options))]
            content["options"] = options
            content["ar_options"] = options
            content["en_options"] = options
            content["question"] = "Select the correct translation: \"{}\"".format(en_word)
            content["en_question"] = "Select the correct translation: \"{}\"".format(en_word)
            content["ar_question"] = "اختر الترجمة الصحيحة: \"{}\"".format(ar_word)
            content["vid"] = vid_id
            
            assert content["answer"] in idx2alp
        else:
            try:
                content = generate_sent_quiz(word)
                idx2alp = ["A", "B", "C", "D"]
                rd.shuffle(content["options"])
                for opt_idx, opt in enumerate(content["options"]):
                    if opt == content["answer"]:
                        content["answer"] = idx2alp[opt_idx]
                        break
                if not content["answer"] in idx2alp:
                    print ("{} answer error".format(word))
                    continue
                for opt_idx in range(len(content["options"])):
                    content["options"][opt_idx] = "{}. {}".format(idx2alp[opt_idx], content["options"][opt_idx])
                
                multi_lingual_quiz = video_processor.translate_zh_quiz(content)
                content["ar_question"] = multi_lingual_quiz["ar_quiz"]["question"]
                # content["ar_options"] = multi_lingual_quiz["ar_quiz"]["options"]
                content["ar_options"] = content["options"]
                content["ar_explanation"] = multi_lingual_quiz["ar_quiz"]["explanation"]
                content["en_question"] = multi_lingual_quiz["en_quiz"]["question"]
                # content["en_options"] = multi_lingual_quiz["en_quiz"]["options"]
                content["en_options"] = content["options"]
                content["en_explanation"] = multi_lingual_quiz["en_quiz"]["explanation"]
                content["vid"] = vid_id
            except Exception as e:
                print (e)
                print ("{} error".format(word))
                time.sleep(5)
        df_list[i].append(json.dumps(content, ensure_ascii=False))
        df_list[i].append(content["vid"])
        fw.write("{}\n".format(json.dumps(content, ensure_ascii=False)))

    df_new = pd.DataFrame(df_list, columns=columns)
    df_new.to_csv(quiz_csv, index=False)
    fw.close()

def gif_to_mp4(gif_path):
    cmd = "/opt/homebrew/Cellar/ffmpeg/7.1_4/bin/ffmpeg -loglevel error -y -i  \"{}\" -vf \"scale=trunc(iw/2)*2:trunc(ih/2)*2\" -c:v libx264 -pix_fmt yuv420p -movflags +faststart -crf 23 \"{}\"".format(gif_path, gif_path.replace(".gif", ".mp4").replace(".GIF", ".mp4"))
    os.system(cmd)
    return gif_path.replace(".gif", ".mp4").replace(".GIF", ".mp4")

@retry(stop_max_attempt_number=3)
def generate_sent_quiz(word):
    json_str = json.dumps({"sentence": "我跟朋友分别时会说再见。", "question": "我跟朋友分别时会说____。", "options": ["再见", "看见", "再一次", "见到"], "answer": "再见", "explanation": "“再见”是人们在分别时常使用的礼貌用语，而“再说、看见、再一次”并不是适合语境的表达，不符合日常用语习惯，所以应选“再见”。"})
    prompt = "请根据下面给出的中文单词进行造句，同时将其变成一个四选一的题目。要注意选项中的单词除了正确选项外，其余要是一个常用的中文词汇，但非常不适合填在句子中。同时要说明选择正确选项的理由。结果用Json格式进行返回，json格式如下{}。中文单词：{}".format(json_str, word)
    resp = call_doubao_pro_32k(prompt)
    content_str = resp.replace("```json", "").replace("```", "")
    print (content_str)
    content = json.loads(content_str)
    return content
    
def modify_videos(emoji_path, word, pinyin, ar_word, merged_audio, audio_dur_dict, out_video_path):
    audio_dur = sum([audio_dur_dict[key] for key in audio_dur_dict.keys()])
    try:
        if emoji_path.endswith(".gif") or emoji_path.endswith(".GIF"):
            emoji_path = gif_to_mp4(emoji_path)
            # cmd = "/opt/homebrew/Cellar/ffmpeg/7.1_4/bin/ffmpeg -loglevel error -y -i  {} -b:v 1M -vf scale=540:-1 {}".format(emoji_path, os.path.join(root_dir, "tmp.mp4"))
            # os.system(cmd)
            video_ori = VideoFileClip(emoji_path)
            if audio_dur > video_ori.duration:
                n_loop = int(audio_dur / video_ori.duration) + 1
                print (n_loop)
            video_ori = video_ori.loop(n=n_loop)
            # os.remove(os.path.join(root_dir, "tmp.mp4"))
            # video_ori = VideoFileClip(os.path.join(root_dir, "tmp_loop.mp4"))
            # os.remove(os.path.join(root_dir, "tmp_loop.mp4"))   
        else:
            fps = 10
            image_num = int(fps * audio_dur)
            images = [emoji_path for i in range(image_num)]
            video_ori = ImageSequenceClip(images, fps=fps)
        # video_no_audio=video_ori.set_audio(None)

        video = video_ori.subclip(0, audio_dur)
        # video_audio = video.audio.volumex(0.05)
        max_val = max(video.size[0], video.size[1])
        if max_val < 540:
            resize_ratio = max(1, 540.0 / float(max_val))
            new_w = resize_ratio * video.w
            if new_w % 2 == 1:
                new_w += 1
            new_h = resize_ratio * video.h
            if new_h % 2 == 1:
                new_h += 1
            video = video.resize(newsize=(new_w, new_h))
        
            
        txt_clip_chinese = TextClip("{}\n{}".format(pinyin, word), fontsize=35, color='white', font='Songti-SC-Black')
        txt_clip_chinese = txt_clip_chinese.set_position("center", relative=True).set_duration(video.duration)

        # print (info["arabic"])
        # print (get_display(info["arabic"]))
        # r2l_arbic = arabic_reshaper.reshape(get_display(info["arabic"]))
        r2l_arbic = get_display(arabic_reshaper.reshape(ar_word))
        # print (r2l_arbic)
        txt_clip_arbic = TextClip(r2l_arbic, fontsize=35, color='white', font='Almarai')
        txt_clip_arbic = txt_clip_arbic.set_position("center", relative=True).set_duration(video.duration)

        txt_w = max(txt_clip_arbic.w, txt_clip_chinese.w)
        chinese_bg_color = ColorClip(size=(txt_w, txt_clip_chinese.h), color=(0, 0, 0), duration=video.duration)
        chinese_bg_color = chinese_bg_color.set_opacity(0.7)
        chinese_text_with_bg = CompositeVideoClip([chinese_bg_color, txt_clip_chinese])
        chinese_text_with_bg = chinese_text_with_bg.set_position(("center", 0.2), relative=True).set_duration(video.duration)
        arbic_text_with_bg = ColorClip(size=(txt_w, txt_clip_arbic.h), color=(0, 0, 0), duration=video.duration)
        arbic_text_with_bg = arbic_text_with_bg.set_opacity(0.7)
        arbic_text_with_bg = CompositeVideoClip([arbic_text_with_bg, txt_clip_arbic])
        
        arbic_h = float(video.size[1] * 0.2 + chinese_text_with_bg.h) / float(video.size[1])
        arbic_text_with_bg = arbic_text_with_bg.set_position(("center", arbic_h), relative=True).set_duration(video.duration)
        video_with_text = CompositeVideoClip([video, chinese_text_with_bg, arbic_text_with_bg])
        audio_clip = AudioFileClip(merged_audio)

        # final_audio = CompositeAudioClip([audio_clip, video_audio])

        video_final = video_with_text.set_audio(audio_clip)
        # video_final.write_videofile(out_video_path, codec='libx264', audio_codec="aac")

        
        # video_final = CompositeVideoClip([video_final, pre_text_clip_with_bg, word_clip_withbg, post_text_clip_withbg])
        video_final.write_videofile(out_video_path, codec='libx264', audio_codec="aac")

        return 1
    except Exception as e:
        print (e)
        print ("{} error".format(word))
        return -1
        
def upload_video(video_srt_file):
    
    df = pd.read_csv(video_srt_file)
    df.dropna(subset=["zh_srt"], axis=0, how="any", inplace=True)



def extract_images_from_excel(hsk_excel, image_dir):
    """
    从Excel文件中提取嵌入的图片/GIF并保存到指定目录
    Args:
        hsk_excel: Excel文件路径
        image_dir: 图片保存目录
    """
    if not os.path.exists(image_dir):
        os.makedirs(image_dir)
    
    wb = load_workbook(hsk_excel)
    ws = wb.active
    
    # 找到word列和image列的索引
    header_row = next(ws.rows)
    word_col = None
    for idx, cell in enumerate(header_row, 1):
        if cell.value == 'word':
            word_col = idx
            break
    
    if word_col is None:
        raise ValueError("找不到'word'列")
    
    # 遍历所有图片
    for image in ws._images:
        row = image.anchor._from.row + 1  # 图片所在行号
        word = ws.cell(row=row, column=word_col).value
        
        if word:
            try:
                # 获取图片数据
                image_data = image._data()
                
                # 判断文件类型并保存
                import io
                import imghdr
                image_type = imghdr.what(None, h=image_data)
                
                if image_type == 'gif':
                    ext = '.gif'
                else:
                    ext = '.png'
                    
                out_path = os.path.join(image_dir, f"{word}{ext}")
                
                with open(out_path, 'wb') as f:
                    f.write(image_data)
                    
            except Exception as e:
                print(f"处理 {word} 的图片时出错: {str(e)}")
                continue

def convert_emojidir_to_csv(hsk_excel, emoji_dir, gif_dir, out_csv):
    gif_dict = dict()
    for item in os.listdir(gif_dir):
        if item.endswith((".gif", ".GIF")):
            word = os.path.splitext(item)[0]  # 去掉文件后缀
            gif_dict[word] = os.path.join(gif_dir, item)

    emoji_dict = dict()
    for item in os.listdir(emoji_dir):
        if item.endswith((".png", ".PNG")):
            word = os.path.splitext(item)[0]  # 去掉文件后缀
            word = word.split('（')[0].split('(')[0]  # 处理中文括号和英文括号
            if word in gif_dict:
                emoji_dict[word] = gif_dict[word]
            else:
                emoji_dict[word] = os.path.join(emoji_dir, item)
    
    ori_df = pd.read_excel(hsk_excel)
    
    # 过滤数据
    ori_df = ori_df.dropna(subset=['word', 'Arabic'])  # 删除word和Arabic为空的行
    ori_df = ori_df[ori_df['tag'].isna()]  # 保留tag为空的行
    
    # 准备结果数据
    result_data = []
    
    for i in range(ori_df.shape[0]):
        word = ori_df.iloc[i]["word"]
        ori_word = word.strip()
        clean_word = word.split('（')[0].split('(')[0]
        if not clean_word in emoji_dict:
            continue
        arabic = ori_df.iloc[i]["Arabic"]
        
        # 解析括号中的等级信息，获取最后一个括号中的内容
        level = 6  # 默认等级
        level_match = re.search(r'[（(][^）)]*[)）](?![^（(]*[)）])', word)  # 匹配最后一个括号
        if level_match:
            level_text = level_match.group(0)[1:-1]  # 去掉括号，只保留内容
            level_dict = {"一级": 1, "二级": 2, "三级": 3, "四级": 4, "五级": 5, "六级": 6}
            if level_text in level_dict:
                level = level_dict[level_text]
        
        if clean_word in emoji_dict:
            result_data.append({
                "原始单词": ori_word,
                "单词": clean_word,
                "emoji": emoji_dict[clean_word],
                "阿语翻译": arabic,
                "level": level
            })
    
    # 创建DataFrame并保存
    result_df = pd.DataFrame(result_data)
    result_df.to_csv(out_csv, index=False)

def outdate_with_csv(create_csv, date):
    df = pd.read_csv(create_csv)
    from create_video import update_video_info
    for i in tqdm(range(df.shape[0])):
        video_id = df.iloc[i]["video_id"]
        update_video_info(video_id, customize="KAU777_{}".format(date))

def convert_charvideos_to_csv(hsk_excel, video_dir, emoji_csv):
    df = pd.read_excel(hsk_excel)
    df.dropna(subset=["word", "Arabic"], how="any", inplace=True)
    infod = dict()
    for i in tqdm(range(df.shape[0])):
        key = df.iloc[i]["word"]
        arabic = df.iloc[i]["Arabic"]
        chinese = df.iloc[i]["chinese"]
        # infod[key] = (chinese, arabic)
        infod[chinese] = (chinese, arabic)
    # 存储结果数据
    result_data = []
    
    for root, dirs, files in os.walk(video_dir):
        for file in files:
            if file.endswith(".mp4"):
                word = file.split(".")[0]
                if word in infod:
                    chinese, arabic = infod[word]
                    result_data.append({
                        "单词": chinese,
                        "FileName": os.path.join(root, file),
                        "阿语翻译": arabic
                    })
    
    # 创建DataFrame并保存为CSV
    result_df = pd.DataFrame(result_data)
    result_df.to_csv(emoji_csv, index=False)

def process_videos_concurrently(video_ids, series_name):
    batch_size = 10
    total_batches = (len(video_ids) + batch_size - 1) // batch_size  # 向上取整
    
    for batch_idx in tqdm(range(0, len(video_ids), batch_size), total=total_batches, desc="处理视频批次"):
        batch_video_ids = video_ids[batch_idx:batch_idx + batch_size]
        
        with ThreadPoolExecutor(max_workers=10) as executor:
            futures = [
                executor.submit(update_video_info, video_id, series_name=series_name)
                for video_id in batch_video_ids
            ]
            # 等待当前批次所有任务完成
            for future in futures:
                future.result()
        
        # 每批次处理完后暂停1秒，避免请求过于频繁
        time.sleep(1)

if __name__ == "__main__":
    # df = pd.read_csv("/Users/tal/work/lingtok_server/video_process/沙特女子Demo/初级汉语/词汇练习/words_refined.csv")
    # df = pd.read_excel("/Users/tal/work/lingtok_server/video_process/沙特女子Demo/初级汉语/真-初级口语1+2/初级口语1/words_chap1.xls")
    # df = pd.read_excel("/Users/tal/work/lingtok_server/video_process/沙特女子Demo/初级汉语/真-初级口语1+2/初级口语1/chap11/words_chap11.xls")
    
    video_processor = VideoProcessor()

    # root_dir = "/Users/tal/work/lingtok_server/video_process/HSK_表情包"
    root_dir = "/Users/tal/work/lingtok_server/video_process/hsk_写字"
    hsk_excel = "/Users/tal/work/lingtok_server/video_process/HSK_表情包/HSK 表情包.xlsx"

    emoji_dir = os.path.join(root_dir, "学中文表情包")
    gif_dir = "/Users/tal/work/lingtok_server/video_process/HSK_表情包/gif"
    # create_tag = "KAU777"
    create_tag = "PNU888"

    audio_dir = os.path.join(root_dir, "audios")
    if not os.path.exists(audio_dir):
        os.makedirs(audio_dir)
    video_dir = os.path.join(root_dir, "videos")
    if not os.path.exists(video_dir):
        os.makedirs(video_dir)
    
    emoji_csv = os.path.join(root_dir, "emoji.csv")
    en_csv = os.path.join(root_dir, "emoji_with_enword.csv")
    quiz_jsonl = os.path.join(root_dir, "words_quiz.jsonl")
    quiz_csv = os.path.join(root_dir, "quiz.csv")
    word_py_csv = os.path.join(root_dir, "words_refined_with_sent_pinyin.csv")
    word_py_ar_csv = os.path.join(root_dir, "words_refined_with_sent_pinyin_ar.csv")
    video_csv = os.path.join(root_dir, "video.csv")
    srt_csv = os.path.join(root_dir, "srt.csv")
    tag_csv = os.path.join(root_dir, "tag.csv")
    vod_csv = os.path.join(root_dir, "vod.csv")
    create_csv = os.path.join(root_dir, "create.csv")

    extract_images = True
    from_charvideos = False

    skip_enword = False
    skip_quiz = False
    skip_py_arword = False
    skip_video = False
    skip_srt = False
    skip_tag = False
    skip_vod = False
    skip_create = False
    skip_series_name = False
    # series_name = "HSK_表情包"
    series_name = "HSK_写字"


    if not extract_images:
        extract_images_from_excel(hsk_excel, emoji_dir)
    
    if from_charvideos:
        # char_video_dir = "/Users/tal/work/lingtok_server/video_process/HSK_表情包/hsk_写字_70"
        char_video_dir = "/Users/tal/work/lingtok_server/video_process/hsk_写字/HSK防抖-979"

        convert_charvideos_to_csv(hsk_excel, char_video_dir, emoji_csv)
    else:
        convert_emojidir_to_csv(hsk_excel, emoji_dir, gif_dir, emoji_csv)
    

    if not skip_enword:
        trans_word_to_en_with_mapcsv(emoji_csv, en_csv)
    
    df = pd.read_csv(en_csv)


    if not skip_quiz:
        generate_trans_quiz(en_csv, quiz_csv, quiz_jsonl)

    if not skip_py_arword:
        add_pinyin(quiz_csv, word_py_csv)
        # trans_word_to_ar(word_py_csv, word_py_ar_csv)
    if not skip_video and not from_charvideos:
        # fj_video_dir = "/Users/tal/work/lingtok_server/video_process/hw/videos/风景视频"
        # video_path_list = [os.path.join(fj_video_dir, item) for item in os.listdir(fj_video_dir)]
        # rd.shuffle(video_path_list)

        df_new = pd.read_csv(word_py_csv)
        # df_new = pd.read_csv("沙特女子Demo/初级汉语/词汇练习/words_refined_with_sent_pinyin_ar_video.csv")
        columns = df_new.columns.to_list()
        columns.append("FileName")
        columns.append("with_sent")
        df_list = df_new.values.tolist()

        for i in tqdm(range(df_new.shape[0])):
            try:
                emoji_path = df_new.iloc[i]["emoji"]
                word = df_new.iloc[i]["单词"]
                py = df_new.iloc[i]["拼音"]
                ar_word = df_new.iloc[i]["阿语翻译"]
                content = json.loads(df_new.iloc[i]["问题"])
                audio_list = [os.path.join(audio_dir, "{}_1.wav".format(word)), os.path.join(audio_dir, "{}_2.wav".format(word)).format(word), os.path.join(audio_dir, "{}_1.wav".format(word))]
                if not os.path.exists(audio_list[1]):
                    generate_wav(word, audio_list[1], voice_type="BV001_streaming", speed=0.3)
                if not os.path.exists(audio_list[0]):
                    generate_wav(word, audio_list[0], voice_type="BV002_streaming", speed=0.3)
                # if not os.path.exists(audio_list[2]):
                #     generate_wav(content["sentence"], audio_list[2], voice_type="BV001_streaming", speed=0.7)
                # if not os.path.exists(audio_list[3]):
                #     generate_wav(content["sentence"], audio_list[3], voice_type="BV002_streaming", speed=0.8)
                merged_wav = os.path.join(audio_dir, "{}_merged.wav".format(word))
                audio_dur_dict = merge_audios(audio_list, merged_wav, sil_dur=1000)
                # rd.shuffle(video_path_list)
                out_video = os.path.join(video_dir, "{}_{}.mp4".format(i, word))
                if not os.path.exists(out_video):
                    mark = modify_videos(emoji_path, word, py, ar_word, merged_wav, audio_dur_dict, out_video)
                    if len(df_list[i]) == 27:
                        df_list[i][-2] = out_video
                    df_list[i][-1] = mark
                else:
                    df_list[i].append(out_video)
                    df_list[i].append(mark)
                
            except Exception as e:
                print (e)
                print ("{} error".format(word))
        
        
        df_video = pd.DataFrame(df_list, columns=columns)
        df_video.to_csv(video_csv, index=False)
    

    if not skip_srt:
        srt_dir = os.path.join(root_dir, "srt_dir")
        if not os.path.exists(srt_dir):
            os.makedirs(srt_dir)
        
        if not from_charvideos:
            df_video = pd.read_csv(video_csv)
        else:
            df_video = pd.read_csv(word_py_csv)
        df_list = df_video.values.tolist()
        columns = df_video.columns.to_list()
        columns += ["zh_srt", "ar_srt", "en_srt", "pinyin_srt"]
        for i in tqdm(range(df_video.shape[0])):
            try:
                word = df_video.iloc[i]["单词"]
                arabic = df_video.iloc[i]["阿语翻译"]
                en_word = df_video.iloc[i]["英语翻译"]
                srt_prefix = "{}_{}".format(i, word)

                srt_res = generate_subtitle_singleword([word], [arabic], [en_word], srt_prefix, srt_dir, 3000)
                df_list[i].append(srt_res["zh_srt"])
                df_list[i].append(srt_res["ar_srt"])
                df_list[i].append(srt_res["en_srt"])
                df_list[i].append(srt_res["pinyin_srt"])
            except Exception as e:
                print (e)
                print ("{} error".format(word))
                time.sleep(5)
        df_srt = pd.DataFrame(df_list, columns=columns)
        df_srt.dropna(subset=["FileName"], inplace=True)
        df_srt.to_csv(srt_csv, index=False)


    if not skip_tag:
        from content_tagger import update_video_info_csv_level
        
        # pnu1_tag_csv = "沙特女子Demo/初级汉语/词汇练习/pnu1_srt_clean_tag.csv"
        update_video_info_csv_level(srt_csv, tag_csv, with_level=False)
        # pnu2_tag_csv = "沙特女子Demo/初级汉语/词汇练习/pnu2_srt_clean_tag.csv"
        # update_video_info_csv_level("沙特女子Demo/初级汉语/词汇练习/pnu2_srt_clean.csv", pnu2_tag_csv)
    
    if not skip_vod:
        from vod_hw_util import upload_hw_withcsv
        # pnu1_vod_csv = "沙特女子Demo/初级汉语/词汇练习/pnu1_srt_clean_tag_vod.csv"
        # upload_hw_withcsv(pnu1_tag_csv, pnu1_vod_csv)
        # pnu2_vod_csv = "沙特女子Demo/初级汉语/词汇练习/pnu2_srt_clean_tag_vod.csv"
        upload_hw_withcsv(tag_csv, vod_csv)

    if not skip_create:
        from create_video import create_with_csv, update_videoinfo_recommender_withcsv
        # pnu1_create_csv = "沙特女子Demo/初级汉语/词汇练习/pnu1_srt_clean_tag_vod_create.csv"
        # create_with_csv("/Users/tal/work/lingtok_server/video_process/沙特女子Demo/初级汉语/词汇练习/words_quiz.jsonl", pnu1_vod_csv, pnu1_create_csv, customize="PNU_1", series_name="PNU_1")
        # update_videoinfo_recommender_withcsv(pnu1_create_csv)

        # pnu2_create_csv = "沙特女子Demo/初级汉语/词汇练习/pnu2_srt_clean_tag_vod_create.csv"
        create_with_csv(quiz_jsonl, vod_csv, create_csv, customize=create_tag, series_name=series_name)
        # update_videoinfo_recommender_withcsv(pnu2_create_csv)
    
    if not skip_series_name:
        from create_video import update_video_info
        import sys
        import os
        sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        from recommender.video_updater import VideoUpdater
        video_updater = VideoUpdater()
        video_updater.update_series_tag_once(series_name, level="初学", tag_list=["科学教育"])
        
        df = pd.read_csv(create_csv)
        # 收集所有有效的video_ids
        video_ids = [vid for vid in df['video_id'].dropna()]
        # 并发处理视频更新
        process_videos_concurrently(video_ids, series_name)

    
